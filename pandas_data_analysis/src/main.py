"""
Module for comparing transaction data with master data and highlighting discrepancies.
"""

import pandas as pd
from pathlib import Path
import logging
from typing import Tuple, Optional
import sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data_processing.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class DataComparator:
    """
    A class to compare transaction data with master data and highlight discrepancies.
    """
    
    def __init__(self, master_path: Path, transaction_path: Path):
        """
        Initialize the DataComparator with file paths.
        
        Args:
            master_path: Path to the master data Excel file
            transaction_path: Path to the transaction data Excel file
        """
        self.master_path = master_path
        self.transaction_path = transaction_path
        self.master_df = None
        self.transaction_df = None
        
    def load_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Load data from Excel files.
        
        Returns:
            Tuple containing master and transaction DataFrames
            
        Raises:
            FileNotFoundError: If input files are not found
            ValueError: If data cannot be loaded properly
        """
        try:
            logger.info(f"Loading master data from {self.master_path}")
            self.master_df = pd.read_excel(self.master_path)
            
            logger.info(f"Loading transaction data from {self.transaction_path}")
            self.transaction_df = pd.read_excel(self.transaction_path)
            
            # Ensure both DataFrames have the same columns in the same order
            common_cols = [col for col in self.master_df.columns if col in self.transaction_df.columns]
            self.master_df = self.master_df[common_cols]
            self.transaction_df = self.transaction_df[common_cols]
            
            logger.info("Data loaded successfully")
            return self.master_df, self.transaction_df
            
        except FileNotFoundError as e:
            logger.error(f"Input file not found: {e}")
            raise
        except Exception as e:
            logger.error(f"Error loading data: {e}")
            raise ValueError(f"Data loading failed: {e}")
    
    def compare_data(self) -> pd.DataFrame:
        """
        Compare transaction data with master data and identify discrepancies.
        
        Returns:
            DataFrame with discrepancies marked
            
        Raises:
            ValueError: If data hasn't been loaded or comparison fails
        """
        if self.master_df is None or self.transaction_df is None:
            raise ValueError("Data not loaded. Call load_data() first.")
            
        try:
            logger.info("Comparing transaction data with master data")
            
            # Create a copy of transaction data to mark discrepancies
            result_df = self.transaction_df.copy()
            
            # Compare each cell with master data
            for col in result_df.columns:
                if col in self.master_df.columns:
                    # Mark discrepancies
                    result_df[f"{col}_Discrepancy"] = result_df[col] != self.master_df[col]
            
            logger.info("Comparison completed successfully")
            return result_df
            
        except Exception as e:
            logger.error(f"Error during data comparison: {e}")
            raise ValueError(f"Data comparison failed: {e}")
    
    def save_comparison_result(self, result_df: pd.DataFrame, output_file: Path):
        """
        Save comparison results to Excel with highlighting.
        
        Args:
            result_df: DataFrame with comparison results
            output_file: Path to save the output Excel file
        """
        try:
            logger.info(f"Saving results to {output_file}")
            
            # Create a new Excel workbook
            wb = Workbook()
            ws = wb.active
            
            # Define styles
            thin_border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
            
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            center_aligned = Alignment(horizontal="center")
            
            # Write DataFrame to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border  # Apply border to all cells
                    
                    # Style header row
                    if r_idx == 1:  # Header row
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_aligned
            
            # Apply highlighting for discrepancies
            for col in result_df.columns:
                if "_Discrepancy" in col:
                    base_col = col.replace("_Discrepancy", "")
                    col_idx = list(result_df.columns).index(base_col) + 1  # +1 for 1-based indexing
                    discrep_col_idx = list(result_df.columns).index(col) + 1
                    
                    for row_idx in range(2, len(result_df) + 2):  # Start from row 2 (skip header)
                        if ws.cell(row=row_idx, column=discrep_col_idx).value:
                            ws.cell(row=row_idx, column=col_idx).fill = red_fill
            
            # Remove discrepancy columns (optional)
            cols_to_delete = [col for col in result_df.columns if "_Discrepancy" in col]
            for col in sorted(cols_to_delete, reverse=True):
                ws.delete_cols(list(result_df.columns).index(col) + 1)
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_file)
            logger.info("Results saved successfully")
            
        except Exception as e:
            logger.error(f"Error saving results: {e}")
            raise

def main():
    """Main execution function."""
    try:
        # Define paths
        input_dir = Path("data/input")
        output_dir = Path("data/output")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        master_file = input_dir / "Master Data.xlsx"
        transaction_file = input_dir / "Transaction Data.xlsx"
        output_file = output_dir / "Transaction_Data_Compared.xlsx"
        
        # Initialize and process data
        comparator = DataComparator(master_file, transaction_file)
        master_df, transaction_df = comparator.load_data()
        
        # Log basic info about the data
        logger.info(f"Master data shape: {master_df.shape}")
        logger.info(f"Transaction data shape: {transaction_df.shape}")
        
        # Compare data
        result_df = comparator.compare_data()
        
        # Save results with highlighting
        comparator.save_comparison_result(result_df, output_file)
        logger.info("Processing completed successfully")
        
    except Exception as e:
        logger.error(f"Processing failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()