"""
Module for comparing transaction data with master data with improved column handling.
"""

import pandas as pd
from pathlib import Path
import logging
from typing import Tuple, Optional, List
import sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data_processing.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class DataComparator:
    """
    A class to compare transaction data with master data with intelligent column handling.
    """
    
    def __init__(self, master_path: Path, transaction_path: Path, key_columns: Optional[List[str]] = None):
        self.master_path = master_path
        self.transaction_path = transaction_path
        self.key_columns = key_columns
        self.master_df = None
        self.transaction_df = None
        self.duplicates_report = {
            'master_duplicates_removed': 0,
            'transaction_duplicates_removed': 0
        }
        
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean the dataframe by removing unnamed columns and converting data types."""
        # Remove unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # Convert common columns to consistent types
        if 'Quantity' in df.columns:
            df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce')
        if 'Part Number' in df.columns:
            df['Part Number'] = df['Part Number'].astype(str)
        if 'Description' in df.columns:
            df['Description'] = df['Description'].astype(str)
        
        return df
    
    def _get_relevant_columns(self) -> List[str]:
        """Get meaningful columns for matching, excluding dates and quantities."""
        common_cols = list(set(self.master_df.columns) & set(self.transaction_df.columns))
        
        # Exclude columns that are likely not good for matching
        exclude = ['Valid From', 'Valid To', 'Quantity']
        candidate_cols = [col for col in common_cols if col not in exclude]
        
        # If user specified key columns, use those that exist
        if self.key_columns:
            return [col for col in self.key_columns if col in candidate_cols]
        
        return candidate_cols or common_cols  # Fall back to all common columns if no candidates
    
    def _align_data(self) -> None:
        """Align data using the best available columns."""
        key_columns = self._get_relevant_columns()
        
        if not key_columns:
            raise ValueError("No suitable columns found for matching records")
        
        logger.info(f"Using key columns for matching: {key_columns}")
        
        # Remove duplicates
        self.master_df = self._remove_duplicates(self.master_df, 'master', key_columns)
        self.transaction_df = self._remove_duplicates(self.transaction_df, 'transaction', key_columns)
        
        # Ensure consistent data types
        for col in key_columns:
            if col in self.master_df.columns and col in self.transaction_df.columns:
                common_type = self._get_common_dtype(self.master_df[col], self.transaction_df[col])
                self.master_df[col] = self.master_df[col].astype(common_type)
                self.transaction_df[col] = self.transaction_df[col].astype(common_type)
        
        # Merge on key columns
        merged = pd.merge(
            self.transaction_df,
            self.master_df,
            on=key_columns,
            how='inner',
            suffixes=('', '_master')
        )
        
        if len(merged) == 0:
            raise ValueError(f"No matching records found using columns: {key_columns}")
        
        # Separate back into transaction and master data
        master_cols = [col for col in self.master_df.columns if col in merged.columns]
        trans_cols = [col for col in self.transaction_df.columns if col in merged.columns]
        
        self.master_df = merged[master_cols]
        self.transaction_df = merged[trans_cols]
        
    def _get_common_dtype(self, s1: pd.Series, s2: pd.Series) -> str:
        """Determine a common data type that can accommodate both series."""
        if s1.dtype == s2.dtype:
            return s1.dtype
        
        # Handle common cases
        if str(s1.dtype) == 'object' or str(s2.dtype) == 'object':
            return 'object'
        if 'float' in str(s1.dtype) or 'float' in str(s2.dtype):
            return 'float64'
        return 'object'
    
    def _remove_duplicates(self, df: pd.DataFrame, df_type: str, key_columns: List[str]) -> pd.DataFrame:
        """Remove duplicate rows based on key columns."""
        initial_count = len(df)
        df = df.drop_duplicates(subset=key_columns, keep='first')
        duplicates_removed = initial_count - len(df)
        self.duplicates_report[f'{df_type}_duplicates_removed'] = duplicates_removed
        
        if duplicates_removed > 0:
            logger.warning(f"Removed {duplicates_removed} duplicates from {df_type} data")
        
        return df
    
    def load_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Load and prepare data for comparison."""
        try:
            logger.info(f"Loading master data from {self.master_path}")
            self.master_df = pd.read_excel(self.master_path)
            logger.info(f"Master data columns: {self.master_df.columns.tolist()}")
            
            logger.info(f"Loading transaction data from {self.transaction_path}")
            self.transaction_df = pd.read_excel(self.transaction_path)
            logger.info(f"Transaction data columns: {self.transaction_df.columns.tolist()}")
            
            # Clean data
            self.master_df = self._clean_data(self.master_df)
            self.transaction_df = self._clean_data(self.transaction_df)
            
            # Align data
            self._align_data()
            
            logger.info(f"Aligned data shapes - Master: {self.master_df.shape}, Transaction: {self.transaction_df.shape}")
            return self.master_df, self.transaction_df
            
        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            if self.master_df is not None:
                logger.info("Master Data Sample:\n" + str(self.master_df.head(3)))
            if self.transaction_df is not None:
                logger.info("Transaction Data Sample:\n" + str(self.transaction_df.head(3)))
            raise ValueError(f"Data loading failed: {str(e)}")
    
    def compare_data(self) -> pd.DataFrame:
        """Compare transaction data with master data."""
        result_df = self.transaction_df.copy()
        
        # Compare each comparable column
        comparable_cols = [col for col in result_df.columns 
                         if col in self.master_df.columns 
                         and not col.endswith('_master')]
        
        for col in comparable_cols:
            result_df[f"{col}_Discrepancy"] = result_df[col] != self.master_df[col]
        
        return result_df
    
    def save_comparison_result(self, result_df: pd.DataFrame, output_file: Path):
        """Save results to Excel with formatting."""
        wb = Workbook()
        ws = wb.active
        
        # Styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:  # Header
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Highlight discrepancies
        for col in result_df.columns:
            if "_Discrepancy" in col:
                base_col = col.replace("_Discrepancy", "")
                col_idx = list(result_df.columns).index(base_col) + 1
                for row_idx in range(2, len(result_df) + 2):
                    if ws.cell(row=row_idx, column=col_idx + len(result_df.columns)).value:
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill
        
        # Remove discrepancy columns
        for col in sorted([c for c in result_df.columns if "_Discrepancy" in c], reverse=True):
            ws.delete_cols(list(result_df.columns).index(col) + 1)
        
        # Add summary sheet
        summary = wb.create_sheet("Summary")
        summary.append(["Comparison Summary"])
        summary.append(["Master Records:", len(self.master_df)])
        summary.append(["Transaction Records:", len(self.transaction_df)])
        summary.append(["Duplicates Removed (Master):", self.duplicates_report['master_duplicates_removed']])
        summary.append(["Duplicates Removed (Transaction):", self.duplicates_report['transaction_duplicates_removed']])
        
        # Auto-size columns
        for sheet in wb:
            for column in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = max_length + 2
        
        wb.save(output_file)

def main():
    """Main execution function."""
    try:
        input_dir = Path("data/input")
        output_dir = Path("data/output")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        master_file = input_dir / "Master Data.xlsx"
        transaction_file = input_dir / "Transaction Data.xlsx"
        output_file = output_dir / "Comparison_Results.xlsx"
        
        # Initialize with specific key columns that exist in your data
        comparator = DataComparator(
            master_file, 
            transaction_file,
            key_columns=["Part Number", "Description"]  # Use these known good columns
        )
        
        master_df, transaction_df = comparator.load_data()
        result_df = comparator.compare_data()
        comparator.save_comparison_result(result_df, output_file)
        
        logger.info(f"Successfully saved results to {output_file}")
    
    except Exception as e:
        logger.error(f"Processing failed: {str(e)}")
        logger.info("Troubleshooting tips:")
        logger.info("1. Check that both files have the expected columns")
        logger.info("2. Verify that 'Part Number' and 'Description' columns contain matching values")
        logger.info("3. Ensure there are no data type conflicts in these columns")
        sys.exit(1)

if __name__ == "__main__":
    main()